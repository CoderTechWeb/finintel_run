from openpyxl import load_workbook
from datetime import datetime, date
import os
import re

HOURS_PER_DAY = 8.0

LEAVE_THRESHOLD = 3
MARGIN_LOW_THRESHOLD = 10
UTILISATION_LOW_THRESHOLD = 80


def clean(v):
    return str(v).strip() if v else ""


def is_date(v):
    return isinstance(v, (datetime, date))


def _is_header_row(row):
    vals = [clean(v).lower() for v in row]
    return "name" in vals and "project" in vals


def _find_col(header, label):
    for j, v in enumerate(header):
        if v and label in clean(v).lower():
            return j
    return None


def _extract_month_label(rows, sheet_name):
    for row in rows[:3]:
        for cell in row:
            if is_date(cell):
                return cell.strftime("%B %Y")
    m = re.search(r"([A-Z]+)['\s]*(\d{2,4})", sheet_name, re.IGNORECASE)
    if m:
        return f"{m.group(1).capitalize()} 20{m.group(2)[-2:]}"
    return sheet_name


def _parse_fortnight(rows, header_idx):
    header = rows[header_idx]
    name_idx = 0
    project_idx = 1

    vacation_idx = _find_col(header, "vacation")
    working_days_idx = _find_col(header, "working days")

    date_cols = [j for j, v in enumerate(header) if is_date(v)]

    billable_idx = None
    actual_idx = None
    sub_header = rows[header_idx + 1] if header_idx + 1 < len(rows) else None
    if sub_header:
        for j, v in enumerate(sub_header):
            lbl = clean(v).lower()
            if "billable" in lbl and billable_idx is None:
                billable_idx = j
            elif "actual" in lbl and actual_idx is None:
                actual_idx = j

    employees = {}
    found_any = False
    for r in rows[header_idx + 2: header_idx + 50]:
        if not r or not r[0]:
            if found_any:
                break
            continue
        name = clean(r[name_idx])
        if not name or name.lower() in ("name", "total", ""):
            if found_any:
                break
            continue

        project = clean(r[project_idx]) if project_idx < len(r) else ""
        if not project or is_date(r[project_idx] if project_idx < len(r) else None):
            break

        vacation = 0
        if vacation_idx is not None and vacation_idx < len(r) and r[vacation_idx]:
            try:
                vacation = int(float(r[vacation_idx]))
            except:
                pass

        working_days = 0
        if working_days_idx is not None and working_days_idx < len(r) and r[working_days_idx]:
            try:
                working_days = int(float(r[working_days_idx]))
            except:
                pass

        billable_hours = 0
        if billable_idx is not None and billable_idx < len(r) and r[billable_idx]:
            try:
                billable_hours = float(r[billable_idx])
            except:
                pass

        actual_hours = 0
        if actual_idx is not None and actual_idx < len(r) and r[actual_idx]:
            try:
                actual_hours = float(r[actual_idx])
            except:
                pass

        approved_idx = _find_col(header, "approved")
        expected_hours = 0
        if approved_idx is not None and approved_idx < len(r) and r[approved_idx]:
            try:
                expected_hours = float(r[approved_idx])
            except:
                pass

        employees[name] = {
            "project": project,
            "actual_hours": actual_hours,
            "billable_hours": billable_hours,
            "expected_hours": expected_hours,
            "vacation_days": vacation,
            "working_days": working_days,
        }
        found_any = True

    return employees


def _parse_summary_section(rows):
    summary = {}
    header_idx = None
    for i, row in enumerate(rows):
        vals = [clean(v).lower() for v in row]
        if "name" in vals and any("total actual" in v or "billing_rate" in v for v in vals):
            header_idx = i
            break

    if header_idx is None:
        return summary

    header = rows[header_idx]
    total_actual_idx = _find_col(header, "total actual")
    max_billable_idx = _find_col(header, "max billable")
    final_billable_idx = _find_col(header, "final billable")
    billing_rate_idx = _find_col(header, "billing_rate")
    cost_rate_idx = _find_col(header, "cost_rate")

    found_any = False
    for r in rows[header_idx + 1: header_idx + 30]:
        if not r or not r[0]:
            if found_any:
                break
            continue
        name = clean(r[0])
        if not name or name.lower() in ("name", "total", ""):
            if found_any:
                break
            continue

        entry = {}
        if total_actual_idx is not None and total_actual_idx < len(r) and r[total_actual_idx]:
            try:
                entry["total_actual_hours"] = float(r[total_actual_idx])
            except:
                pass
        if max_billable_idx is not None and max_billable_idx < len(r) and r[max_billable_idx]:
            try:
                entry["max_billable_hours"] = float(r[max_billable_idx])
            except:
                pass
        if final_billable_idx is not None and final_billable_idx < len(r) and r[final_billable_idx]:
            try:
                entry["final_billable_hours"] = float(r[final_billable_idx])
            except:
                pass
        if billing_rate_idx is not None and billing_rate_idx < len(r) and r[billing_rate_idx]:
            try:
                entry["billing_rate"] = float(r[billing_rate_idx])
            except:
                pass
        if cost_rate_idx is not None and cost_rate_idx < len(r) and r[cost_rate_idx]:
            try:
                entry["cost_rate"] = float(r[cost_rate_idx])
            except:
                pass

        summary[name] = entry
        found_any = True

    return summary


def _match_name(name, lookup):
    if name in lookup:
        return lookup[name]
    for key, val in lookup.items():
        if name.lower() in key.lower() or key.lower() in name.lower():
            return val
    return None


def _build_employee_record(name, data, month_label):
    actual_hours = data.get("actual_hours", 0)
    billable_hours = data.get("billable_hours", actual_hours)
    expected_hours = data.get("expected_hours", 0)
    working_days = data.get("working_days", 0)
    vacation_days = data.get("vacation_days", 0)
    effective_working_days = working_days - vacation_days
    leave_hours = vacation_days * HOURS_PER_DAY

    billing_rate = data.get("billing_rate", 0)
    cost_rate = data.get("cost_rate", 0)

    revenue = round(billable_hours * billing_rate, 2) if billing_rate else 0
    cost = round((actual_hours + leave_hours) * cost_rate, 2) if cost_rate else 0
    profit = round(revenue - cost, 2)
    margin_pct = round((profit / revenue) * 100, 2) if revenue > 0 else (0 if revenue == 0 and profit == 0 else -100)
    utilisation_pct = round((actual_hours / expected_hours) * 100, 2) if expected_hours > 0 else 0

    loss_reasons = []
    validation_flags = []

    if profit < 0:
        loss_reasons.append("NEGATIVE_MARGIN")
    if vacation_days >= LEAVE_THRESHOLD:
        loss_reasons.append("HIGH_LEAVE_IMPACT")

    if margin_pct < MARGIN_LOW_THRESHOLD and revenue > 0:
        validation_flags.append("LOW_MARGIN")
    if profit < 0:
        validation_flags.append("NEGATIVE_MARGIN")
    if vacation_days >= LEAVE_THRESHOLD:
        validation_flags.append("HIGH_LEAVE")
    if utilisation_pct < UTILISATION_LOW_THRESHOLD:
        validation_flags.append("LOW_UTILISATION")

    return {
        "employee": name,
        "project": data.get("project", ""),
        "month": month_label,
        "working_days": working_days,
        "vacation_days": vacation_days,
        "effective_working_days": effective_working_days,
        "actual_hours": actual_hours,
        "billable_hours": billable_hours,
        "expected_hours": expected_hours,
        "leave_hours": leave_hours,
        "billing_rate": billing_rate,
        "cost_rate": cost_rate,
        "revenue": revenue,
        "cost": cost,
        "profit": profit,
        "margin_pct": margin_pct,
        "utilisation_pct": utilisation_pct,
        "loss_reasons": loss_reasons,
        "validation_flags": validation_flags,
        "is_profitable": profit >= 0,
    }


def _match_target_month(sheet_name, target_month):
    if not target_month:
        return True
    target = target_month.upper().replace("-", "").replace(" ", "").replace("'", "")
    sheet = sheet_name.upper().replace("-", "").replace(" ", "").replace("'", "")
    return target in sheet


def parse_timesheet(filepath, target_month=None):
    wb = load_workbook(filepath, data_only=True)

    result = {
        "file": os.path.basename(filepath),
        "sheets": {}
    }

    for sheet_name in wb.sheetnames:
        if target_month and not _match_target_month(sheet_name, target_month):
            continue
        ws = wb[sheet_name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]

        if len(rows) < 4:
            continue

        month_label = _extract_month_label(rows, sheet_name)

        header_indices = [i for i, row in enumerate(rows) if _is_header_row(row)]
        if not header_indices:
            continue

        fortnight_headers = [i for i in header_indices
                             if _find_col(rows[i], "approved") is not None or
                             any(is_date(v) for v in rows[i])]

        merged = {}
        for h_idx in fortnight_headers:
            fort_data = _parse_fortnight(rows, h_idx)
            for name, data in fort_data.items():
                if name not in merged:
                    merged[name] = {
                        "project": data["project"],
                        "actual_hours": 0,
                        "billable_hours": 0,
                        "expected_hours": 0,
                        "vacation_days": 0,
                        "working_days": 0,
                    }
                merged[name]["actual_hours"] += data["actual_hours"]
                merged[name]["billable_hours"] += data["billable_hours"]
                merged[name]["expected_hours"] += data["expected_hours"]
                merged[name]["vacation_days"] += data["vacation_days"]
                merged[name]["working_days"] += data["working_days"]

        summary_data = _parse_summary_section(rows)

        for name in merged:
            sm = _match_name(name, summary_data)
            if sm:
                merged[name]["billing_rate"] = sm.get("billing_rate", 0)
                merged[name]["cost_rate"] = sm.get("cost_rate", 0)
                if sm.get("total_actual_hours"):
                    merged[name]["actual_hours"] = sm["total_actual_hours"]
                if sm.get("final_billable_hours"):
                    merged[name]["billable_hours"] = sm["final_billable_hours"]
                if sm.get("max_billable_hours"):
                    merged[name]["expected_hours"] = sm["max_billable_hours"]
            else:
                merged[name].setdefault("billing_rate", 0)
                merged[name].setdefault("cost_rate", 0)

        if not merged:
            continue

        employees = [
            _build_employee_record(name, data, month_label)
            for name, data in merged.items()
        ]

        projects = {}
        for emp in employees:
            proj = emp["project"] or "Unknown"
            if proj not in projects:
                projects[proj] = {"revenue": 0, "cost": 0, "profit": 0, "employees": 0}
            projects[proj]["revenue"] += emp["revenue"]
            projects[proj]["cost"] += emp["cost"]
            projects[proj]["profit"] += emp["profit"]
            projects[proj]["employees"] += 1
        for p in projects.values():
            p["revenue"] = round(p["revenue"], 2)
            p["cost"] = round(p["cost"], 2)
            p["profit"] = round(p["profit"], 2)

        total_revenue = round(sum(e["revenue"] for e in employees), 2)
        total_cost = round(sum(e["cost"] for e in employees), 2)
        total_profit = round(sum(e["profit"] for e in employees), 2)
        avg_margin = round((total_profit / total_revenue) * 100, 2) if total_revenue > 0 else 0

        sorted_by_profit = sorted(employees, key=lambda e: e["profit"], reverse=True)
        top_performers = [{"employee": e["employee"], "profit": e["profit"]}
                          for e in sorted_by_profit if e["profit"] > 0][:3]
        low_performers = [{"employee": e["employee"], "profit": e["profit"]}
                          for e in sorted_by_profit if e["profit"] <= 0]

        risks = []
        for e in employees:
            if e["profit"] < 0:
                risks.append({"employee": e["employee"], "issue": "LOSS_MAKING"})
            if e["vacation_days"] >= LEAVE_THRESHOLD:
                risks.append({"employee": e["employee"], "issue": "HIGH_LEAVE"})
            if e["utilisation_pct"] < UTILISATION_LOW_THRESHOLD:
                risks.append({"employee": e["employee"], "issue": "LOW_UTILISATION"})

        result["sheets"][sheet_name] = {
            "template": "proofpoint",
            "summary": {
                "total_employees": len(employees),
                "total_revenue": total_revenue,
                "total_cost": total_cost,
                "total_profit": total_profit,
                "avg_margin_pct": avg_margin,
                "total_actual_hours": round(sum(e["actual_hours"] for e in employees), 2),
                "total_billable_hours": round(sum(e["billable_hours"] for e in employees), 2),
                "total_working_days": sum(e["working_days"] for e in employees),
            },
            "projects": projects,
            "employees": employees,
            "top_performers": top_performers,
            "low_performers": low_performers,
            "risks": risks,
        }

    total_rev = sum(s["summary"]["total_revenue"] for s in result["sheets"].values())
    total_cst = sum(s["summary"]["total_cost"] for s in result["sheets"].values())
    total_pft = sum(s["summary"]["total_profit"] for s in result["sheets"].values())

    result["overall_summary"] = {
        "total_revenue": round(total_rev, 2),
        "total_cost": round(total_cst, 2),
        "total_profit": round(total_pft, 2),
        "avg_margin_pct": round((total_pft / total_rev) * 100, 2) if total_rev > 0 else 0,
    }

    return result